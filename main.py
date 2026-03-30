import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from database import connect_db
from category import create_category, CATEGORIES
from auth import show_login
import pandas as pd
from file_handler import upload_file
from datetime import datetime
from collections import defaultdict
import locale
from decimal import Decimal
from openpyxl.styles import Font, Border, Side, Alignment
import os

conn = connect_db()
if conn:
    cursor = conn.cursor()

# =================================================================== AHSP ===================================================================
def ahsp(frame_container, page_name):
    if page_name == "Input Data Barang & Material":
        upload_button = tk.Button(frame_container, text="AHSP Baru", command=lambda: upload_file(), font=("Arial", 12, "bold"), bg="blue", fg="white")
        upload_button.pack(pady=10)
        
        back_button = tk.Button(frame_container, text="Kembali", command=main_menu, font=("Arial", 10, "bold"), bg="red", fg="white", padx=10, pady=5)
        back_button.pack(pady=10)
            
    elif page_name == "Lihat Data Barang & Material":
        show_ahsp_page(frame_container)
        
def format_rupiah(amount):
    return f"Rp {amount:,.0f}".replace(",", ".")

def show_ahsp_page(frame_container):
    cursor = conn.cursor()

    query = """
        SELECT k.kode_kelompok_barang, k.uraian_kelompok_barang, 
               d.kode_barang, d.uraian_barang, d.spesifikasi, d.satuan, d.harga_satuan, d.tanggal_input_data
        FROM data_barang d
        JOIN kelompok_barang k ON d.kode_kelompok_barang = k.kode_kelompok_barang
        ORDER BY k.kode_kelompok_barang
    """
    cursor.execute(query)
    rows = cursor.fetchall()
    original_data = rows.copy()

    uraian_kelompok_dict = {row[0]: row[1] for row in rows}  
    uraian_kelompok_values = ["Semua"] + list(uraian_kelompok_dict.values())

    top_frame = tk.Frame(frame_container)
    top_frame.pack(fill="x", padx=10, pady=5)

    combo_filter = ttk.Combobox(top_frame, values=uraian_kelompok_values, state="readonly")
    combo_filter.current(0)  
    combo_filter.pack(side="left", padx=5)

    search_label = tk.Label(top_frame, text="Search:")
    search_label.pack(side="right", padx=5)

    search_var = tk.StringVar()
    search_entry = tk.Entry(top_frame, textvariable=search_var)
    search_entry.pack(side="right", padx=5)

    style = ttk.Style()
    style.configure("Treeview.Heading", font=("Arial", 10, "bold"))

    columns = ("Kode Kelompok", "Uraian Kelompok", "Kode Barang", "Uraian Barang", "Spesifikasi", "Satuan", "Harga Satuan", "Tanggal Input AHSP")
    table = ttk.Treeview(frame_container, columns=columns, show="headings", style="Treeview")

    for col in columns:
        table.heading(col, text=col)
        table.column(col, anchor="w", width=150)

    table.pack(fill="both", expand=True)

    total_label = tk.Label(frame_container, text="", font=("Arial", 12, "bold"), bg="white", fg="black", pady=10)
    total_label.pack(fill="x", pady=5)

    def update_table():
        filter_value = combo_filter.get()
        search_text = search_var.get().lower()
        table.delete(*table.get_children())  

        for row in original_data:
            formatted_date = row[-1].strftime('%d-%m-%Y %H:%M:%S') if isinstance(row[-1], datetime) else str(row[-1])
            formatted_row = row[:-2] + (format_rupiah(row[-2]), formatted_date)

            if filter_value != "Semua" and row[1] != filter_value:
                continue

            if search_text not in " ".join(str(cell).lower() for cell in row):
                continue

            table.insert("", "end", values=formatted_row)

        filtered_rows = [row for row in original_data if (filter_value == "Semua" or row[1] == filter_value) and search_text in " ".join(str(cell).lower() for cell in row)]
        total_label.config(text=f"Total Data: {len(filtered_rows)} Data & Jumlah Uraian Kelompok Barang: {len(set(r[1] for r in filtered_rows))} Data")

    update_table()

    combo_filter.bind("<<ComboboxSelected>>", lambda e: update_table())  
    search_entry.bind("<KeyRelease>", lambda e: update_table())  

    def go_back():
        for widget in frame_container.winfo_children():
            widget.destroy()
        main_menu()

    back_button = tk.Button(frame_container, text="Kembali", command=go_back, font=("Arial", 10, "bold"), bg="red", fg="white", padx=10, pady=5)
    back_button.pack(pady=10)
        
# =================================================================== RAB ===================================================================
def get_submenus_from_db(table_name):
    try:
        cursor.execute(f"SELECT id_kategori_pekerjaan, jenis_pekerjaan FROM {table_name} ORDER BY id_kategori_pekerjaan")
        submenus = [(row[0], row[1]) for row in cursor.fetchall()]
        return submenus
    except Exception as e:
        print(f"Error: {e}")
        messagebox.showerror("Error", f"Gagal mengambil data submenus dari {table_name}: {e}")
        return []

def get_ahsp_options(jenis_pekerjaan=None, id_kategori_pekerjaan=None):
    try:
        locale.setlocale(locale.LC_ALL, 'id_ID.UTF-8')  
        
        query = """
        SELECT id_kelompok_pekerjaan, kelompok_pekerjaan, nama_pekerjaan, kategori, uraian, satuan, koefisien, harga_satuan, total_harga, jumlah_total, jenis_pekerjaan
        FROM data_ahsp
        """
        
        if id_kategori_pekerjaan:
            placeholders = ', '.join(['%s'] * len(id_kategori_pekerjaan))  
            query += f" WHERE id_kategori_pekerjaan IN ({placeholders})"
        
        query += " ORDER BY id_kelompok_pekerjaan, kelompok_pekerjaan, id_nama_pekerjaan;"
        
        cursor.execute(query, tuple(id_kategori_pekerjaan))  
        
        results = cursor.fetchall()

        ahsp_dict = defaultdict(lambda: defaultdict(dict))  

        for row in results:
            id_kelompok_pekerjaan = row[0]
            kelompok_pekerjaan = row[1]
            nama_pekerjaan = row[2]
            harga_satuan = float(row[9])  
            jenis = row[10]
            
            if jenis_pekerjaan is None or jenis == jenis_pekerjaan:
                ahsp_dict[kelompok_pekerjaan][nama_pekerjaan] = harga_satuan

        return ahsp_dict
    except Exception as e:
        print(f"Error: {e}")
        messagebox.showerror("Error", f"Gagal mengambil data dari tabel AHSP: {e}")
        return {}

def export_to_excel(table, page_name):
    try:
        data = []
        columns = ("NO", "JENIS PEKERJAAN", "URAIAN PEKERJAAN", "VOLUME", "SATUAN", "HARGA SATUAN (Rp)", "JUMLAH (Rp)")

        for item in table.get_children():
            values = table.item(item, "values")
            data.append(values)

        if not data:
            messagebox.showwarning("Peringatan", "Tidak ada data untuk diekspor!")
            return

        df = pd.DataFrame(data, columns=columns)

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Simpan File Excel"
        )

        if file_path:
            writer = pd.ExcelWriter(file_path, engine="openpyxl")
            df.to_excel(writer, index=False, sheet_name="RAB Data", startrow=2)
            workbook = writer.book
            worksheet = writer.sheets["RAB Data"]

            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = page_name.upper()
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal="center")
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            bold_font = Font(bold=True)
            center_alignment = Alignment(horizontal="center")

            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=len(columns)):
                for cell in row:
                    cell.border = thin_border

            for row_num in range(3, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row_num, column=6).value
                if cell_value in ["SUB TOTAL (Rp)", "GRAND TOTAL (Rp)"]:
                    merge_range = f"A{row_num}:E{row_num}"
                    worksheet.merge_cells(merge_range)

                    merged_cell = worksheet.cell(row=row_num, column=1)
                    merged_cell.value = cell_value
                    merged_cell.font = bold_font
                    merged_cell.alignment = center_alignment

                    worksheet.cell(row=row_num, column=6).value = ""

            workbook.save(file_path)
            workbook.close()

            messagebox.showinfo("Sukses", f"Data berhasil diekspor ke {file_path}")

    except Exception as e:
        messagebox.showerror("Error", f"Gagal mengekspor data: {e}")

def rab(frame_container, page_name):
    for widget in frame_container.winfo_children():
        widget.destroy()
    
    kategori_mapping = {
        "Rancangan Anggaran Biaya Paving Cetak": ["KT-PVK001"],
        "Rancangan Anggaran Biaya Paving Cor": ["KT-PVR001"],
        "Rancangan Anggaran Biaya Drainase Cor": ["KT-DRC001"],
        "Rancangan Anggaran Biaya Drainase Buis Beton": ["KT-DRB001"],
        "Rancangan Anggaran Biaya Lantai Rumah": ["KT-RLT001"],
        "Rancangan Anggaran Biaya Dinding Rumah": ["KT-RDG001"],
        "Rancangan Anggaran Biaya Atap Rumah": ["KT-RAT001"],
        "Rancangan Anggaran Biaya Rabat Beton": ["KT-RBT001"],
        "Rancangan Anggaran Biaya Plengsengan": ["KT-PLS001"]
    }
    
    id_kategori_pekerjaan = kategori_mapping.get(page_name, [])
    submenus = get_submenus_from_db("jenis_pekerjaan")
    submenus = [submenu[1] for submenu in submenus if submenu[0] in id_kategori_pekerjaan]
    
    left_frame = tk.Frame(frame_container, bg="white")
    left_frame.pack(side="left", padx=10, pady=10, fill="both", expand=True)
    
    right_frame = tk.Frame(frame_container, bg="white")
    right_frame.pack(side="right", padx=10, pady=10, fill="both", expand=True)
    
    columns = ("NO", "JENIS PEKERJAAN", "URAIAN PEKERJAAN", "VOLUME", "SATUAN", "HARGA SATUAN (Rp)", "JUMLAH (Rp)")
    table = ttk.Treeview(right_frame, columns=columns, show="headings")
    
    for col in columns:
        table.heading(col, text=col)
        table.column(col, width=100, anchor="center")
    
    table.pack(side="top", fill="both", expand=True)
    
    button_frame = tk.Frame(right_frame, bg="white")
    button_frame.pack(side="bottom", fill="x")
    
    def delete_selected():
        selected_item = table.selection()
        if selected_item:
            table.delete(selected_item)
    
    delete_button = tk.Button(button_frame, text="Hapus Baris Terpilih", command=delete_selected, bg="red", fg="white")
    delete_button.pack(pady=10, fill="x")
    
    def add_row(no, jenis, uraian, volume, satuan, harga_satuan, jumlah):
        table.insert("", "end", values=(no, jenis, uraian, volume, satuan, harga_satuan, jumlah))
    
    for submenu in submenus:
        btn = tk.Button(left_frame, text=submenu, command=lambda s=submenu: open_input_window(s, table, id_kategori_pekerjaan))
        btn.pack(pady=5, fill="x")
    
    back_to_main_button = tk.Button(left_frame, text="Kembali ke Menu Utama", command=main_menu, font=("Arial", 12, "bold"), bg="red", fg="white")
    back_to_main_button.pack(pady=10, fill="x")
    
    export_button = tk.Button(right_frame, text="Download Excel", command=lambda: export_to_excel(table, page_name))
    export_button.pack(side="bottom", pady=15)
    export_button.place(relx=0.5, rely=0.90, anchor="center")

def open_input_window(submenu, table, id_kategori_pekerjaan):
    input_window = tk.Toplevel()
    input_window.title(f"Input Data - {submenu}")
    input_window.geometry("450x350")
    input_window.configure(bg="#f0f0f0")

    frame = ttk.LabelFrame(input_window, text="Masukkan Data", padding=(10, 5))
    frame.pack(fill="both", expand=True, padx=10, pady=10)

    ttk.Label(frame, text="Kelompok Pekerjaan:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    ttk.Label(frame, text="List AHSP:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
    ttk.Label(frame, text="Volume:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
    ttk.Label(frame, text="Satuan:").grid(row=3, column=0, padx=5, pady=5, sticky="w")

    jenis_pekerjaan = submenu
    ahsp_dict = get_ahsp_options(jenis_pekerjaan, id_kategori_pekerjaan)

    if not isinstance(ahsp_dict, dict):
        print("Kesalahan: ahsp_dict bukan dictionary")
        messagebox.showerror("Error", "Data tidak valid!")
        return

    if jenis_pekerjaan == "Pekerjaan Persiapan":
        ahsp_dict["Pekerjaan Air Kerja"] = {"Pekerjaan Air Kerja": 0}
        ahsp_dict["Pekerjaan Listrik Kerja"] = {"Pekerjaan Listrik Kerja": 0}

    kelompok_var = tk.StringVar()
    kategori_var = tk.StringVar()
    satuan_var = tk.StringVar()

    kelompok_dropdown = ttk.Combobox(frame, textvariable=kelompok_var, values=list(ahsp_dict.keys()), state="readonly", width=35)
    kelompok_dropdown.grid(row=0, column=1, padx=5, pady=5)

    ahsp_dropdown = ttk.Combobox(frame, textvariable=kategori_var, state="readonly", width=35)
    ahsp_dropdown.grid(row=1, column=1, padx=5, pady=5)

    volume_entry = ttk.Entry(frame, width=37)
    volume_entry.grid(row=2, column=1, padx=5, pady=5)

    satuan_dropdown = ttk.Combobox(frame, textvariable=satuan_var, values=["m", "m2", "m3", "gram", "kg", "buah", "lumpsum"], state="readonly", width=35)
    satuan_dropdown.grid(row=3, column=1, padx=5, pady=5)

    def update_ahsp_options(event):
        selected_kelompok = kelompok_var.get()
        if selected_kelompok in ahsp_dict:
            ahsp_dropdown["values"] = list(ahsp_dict[selected_kelompok].keys())
        else:
            ahsp_dropdown["values"] = []

    kelompok_dropdown.bind("<<ComboboxSelected>>", update_ahsp_options)
    
    notes_label = ttk.Label(frame, text="* Gunakan titik (.) untuk desimal pada input volume.\n* Pastikan semua kelompok pekerjaan telah diinput.", 
                            foreground="red", background="#f0f0f0", anchor="w", justify="left")
    notes_label.grid(row=4, column=0, columnspan=2, padx=5, pady=(5, 15), sticky="w")

    def add_to_table():
        kelompok = kelompok_var.get()
        ahsp = kategori_var.get()
        volume = volume_entry.get()
        satuan = satuan_var.get()

        if not (kelompok and ahsp and volume and satuan):
            messagebox.showerror("Error", "Semua field harus diisi!")
            return

        try:
            volume = float(volume)
            harga_satuan = ahsp_dict[kelompok][ahsp]
            jumlah = volume * harga_satuan

            existing_items = table.get_children()
            for item in existing_items:
                values = table.item(item, "values")
                existing_jenis_pekerjaan = values[1]  
                existing_kelompok = values[2]  

                if existing_jenis_pekerjaan == jenis_pekerjaan and existing_kelompok == kelompok:
                    messagebox.showerror("Error", f"Kelompok pekerjaan '{kelompok}' sudah ditambahkan!")
                    return

            table.insert("", "end", values=(
                len(existing_items) + 1,
                jenis_pekerjaan,
                kelompok,
                volume,
                satuan,
                locale.currency(harga_satuan, grouping=True),
                locale.currency(jumlah, grouping=True)
            ))

            kelompok_terdaftar = {table.item(child, "values")[2] for child in table.get_children() if table.item(child, "values")[1] == jenis_pekerjaan}
            semua_kelompok = set(ahsp_dict.keys())

            if kelompok_terdaftar == semua_kelompok:
                subtotal = sum(
                    float(table.item(child, "values")[6].replace("Rp", "").replace(".", "").replace(",", "."))
                    for child in table.get_children()
                    if table.item(child, "values")[1] == jenis_pekerjaan and table.item(child, "values")[2] != "SUB TOTAL (Rp)"
                )

                table.insert("", "end", values=(
                    "",
                    "",
                    "",
                    "",
                    "",
                    "SUB TOTAL (Rp)",
                    locale.currency(subtotal, grouping=True)
                ), tags=("subtotal",))

                hitung_grand_total()

            input_window.destroy()

        except ValueError:
            messagebox.showerror("Error", "Volume harus berupa angka!")

        table.tag_configure("subtotal", font=("Arial", 10, "bold"), background="#85C4F2")

    def hitung_grand_total():
        grand_total = 0
        sub_total_items = []

        for item in table.get_children():
            values = table.item(item, "values")
            if values[5] == "SUB TOTAL (Rp)":
                subtotal_value = float(values[6].replace("Rp", "").replace(".", "").replace(",", "."))
                grand_total += subtotal_value
                sub_total_items.append(item)

        for item in table.get_children():
            values = table.item(item, "values")
            if values[5] == "GRAND TOTAL (Rp)":
                table.delete(item)
                break

        table.insert("", "end", values=(
            "",
            "",
            "",
            "",
            "",
            "GRAND TOTAL (Rp)",
            locale.currency(grand_total, grouping=True)
        ), tags=("grand_total",))

        table.tag_configure("grand_total", font=("Arial", 11, "bold"), background="#85C4F2")

    button_frame = ttk.Frame(frame)
    button_frame.grid(row=5, column=0, columnspan=2, pady=10)

    ttk.Button(button_frame, text="Batal", command=input_window.destroy, style="TButton").pack(side="left", padx=5)
    ttk.Button(button_frame, text="Tambah", command=add_to_table, style="TButton").pack(side="left", padx=5)

# def kebutuhan(frame_container, page_name):
#     if page_name == "Kebutuhan Drainase":
#         submenus = get_submenus_from_db("kebutuhan")  # Ambil semua data dari database
        
#         # **Filter hanya yang memiliki id_kategori_pekerjaan = 'DRN001'**
#         submenus = [submenu[1] for submenu in submenus if submenu[0] == 'CDRN001']

#         subtotals = {}
#         completed_tasks = {submenu: [] for submenu in submenus}

#         left_frame = tk.Frame(frame_container, bg="white")
#         left_frame.pack(side="left", padx=10, pady=10, fill="both", expand=True)
        
#         columns = ("No.", "Uraian Pekerjaan", "Panjang", "Lebar Atas", "Lebar Bawah", "Tinggi 1", "Tinggi 2", "Volume", "Detail Kebutuhan")
#         table = ttk.Treeview(frame_container, columns=columns, show="headings")

#         for col in columns:
#             table.heading(col, text=col)

#         table.column("No.", width=50, anchor="center")
#         table.column("Uraian Pekerjaan", width=200, anchor="center")
#         table.column("Panjang", width=100, anchor="center")
#         table.column("Lebar Atas", width=100, anchor="center")
#         table.column("Lebar Bawah", width=100, anchor="center")
#         table.column("Tinggi 1", width=100, anchor="center")
#         table.column("Tinggi 2", width=100, anchor="center")
#         table.column("Volume", width=150, anchor="center")
#         table.column("Detail Kebutuhan", width=150, anchor="center")

#         table.pack(side="right", fill="both", expand=True)
        
#         # for submenu in submenus:
#         #     btn = tk.Button(
#         #         left_frame, text=submenu, font=("Arial", 12, "bold"), bg="blue", fg="white", padx=20, pady=10, borderwidth=0,
#         #         command=lambda s=submenu: (
#         #             input_window(root, table, page_name, s, subtotals, completed_tasks)
#         #             if s in submenus else messagebox.showinfo("Info", f"Menu {s} dipilih")
#         #         )
#         #     )
#         #     btn.pack(pady=5, fill="x")
            
#         # Tambahkan tombol kembali ke main menu di bawah list submenus
#         back_to_main_button = tk.Button(left_frame, text="Kembali ke Menu Utama", command=main_menu, font=("Arial", 12, "bold"), bg="red", fg="white")
#         back_to_main_button.pack(pady=10, fill="x")

def detail_menu(page_name):
    for widget in root.winfo_children():
        widget.destroy()
    
    header = tk.Label(root, text=page_name, font=("Arial", 16, "bold"), bg="#1E88E5", fg="white", padx=20, pady=10)
    header.pack(fill="x")
    
    frame_container = tk.Frame(root, bg="white")
    frame_container.pack(fill="both", expand=True, padx=20, pady=20)
    
    if page_name in ["Input Data Barang & Material", "Lihat Data Barang & Material"]:
        ahsp(frame_container, page_name)

    if page_name in ["Rancangan Anggaran Biaya Paving Cetak", "Rancangan Anggaran Biaya Paving Cor", 
                     "Rancangan Anggaran Biaya Drainase Cor", "Rancangan Anggaran Biaya Drainase Buis Beton", 
                     "Rancangan Anggaran Biaya Lantai Rumah", "Rancangan Anggaran Biaya Dinding Rumah", "Rancangan Anggaran Biaya Atap Rumah",
                     "Rancangan Anggaran Biaya Rabat Beton",
                     "Rancangan Anggaran Biaya Plengsengan"]:
        rab(frame_container, page_name)
                
    # if "KALKULATOR KEBUTUHAN" in CATEGORIES:
    #     kebutuhan(frame_container, page_name)
    
def main_menu():
    for widget in root.winfo_children():
        widget.destroy()

    header = tk.Label(root, text="SISTEM RANCANGAN ANGGARAN DESA JUNREJO", font=("Arial", 16, "bold"), bg="#1E88E5", fg="white", padx=20, pady=10)
    header.pack(fill="x")

    frame_main = tk.Frame(root, bg="white")
    frame_main.pack(pady=20, padx=20, fill="both", expand=True)

    for title, items in CATEGORIES.items():
        create_category(frame_main, title, items, detail_menu)

    logout_button = tk.Button(root, text="LOG OUT", font=("Arial", 12, "bold"), bg="red", fg="white", padx=20, pady=10, borderwidth=0, command=lambda: show_login(root, main_menu))
    logout_button.pack(side="bottom", pady=20)

root = tk.Tk()
root.title("SIRADJU")

width = root.winfo_screenwidth() 
height = root.winfo_screenheight()
root.geometry("%dx%d" % (width, height))
root.configure(bg="white")

detail_menu("Input File Excel")
show_login(root, main_menu)
root.mainloop()
